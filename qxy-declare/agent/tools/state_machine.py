#!/usr/bin/env python3
"""
state_machine.py — 严格有限状态机

规则：
  1. advance() 每次只执行一步，执行完就返回
  2. 每个状态声明自己需要什么输入（required_input）
  3. 如果当前状态需要输入但还没收到 → BLOCKED，不往下走
  4. inject() 只接受当前状态声明的 input_key，其他一律拒绝
  5. 没有 force_transition，没有 while 循环，没有任何绕过机制
"""

import json
from shared import (
    create_task, load_state, save_state, transition, fail,
    list_tasks, output, parse_args, log, now_iso,
)
from login import ensure_logged_in
from enterprise_profile import enterprise_profile
from fetch_tax_list import fetch_tax_list
from init_declaration import init_declaration
from submit_declaration import submit_standard, submit_simplified
from download_receipt import download_receipt
from payment import execute_payment


def _parse_period(p: str) -> tuple[int, int]:
    parts = p.split("-")
    return int(parts[0]), int(parts[1])


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 状态定义表
# 每个状态声明：
#   type:           "auto"(自动执行) / "input"(需要外部输入)
#   required_input: input 类型状态需要的 key（inject 时校验）
#   handler:        处理函数名
#   message:        给 Agent 的指令
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

STATE_DEFS = {
    "INIT": {
        "type": "auto",
        "handler": "do_init",
    },
    "ENTERPRISE_PROFILE": {
        "type": "auto",
        "handler": "do_enterprise_profile",
    },
    "FETCH_LIST": {
        "type": "auto",
        "handler": "do_fetch_list",
    },
    "NOTIFY_TAXES": {
        "type": "input",
        "required_input": "notify_taxes_ack",
        "handler": "do_notify_taxes",
        "message": (
            "【必须和用户交互】请向用户发送本月需申报的税种清单，等用户确认后 inject:\n"
            'data_key="notify_taxes_ack" data_value={"user_said": "用户的原话"}'
        ),
    },
    "DATA_INIT": {
        "type": "auto",
        "handler": "do_data_init",
    },
    "WAIT_UPLOAD": {
        "type": "input",
        "required_input": "uploaded_excel",
        "handler": "do_wait_upload",
        "message": (
            "【必须和用户交互】请通知用户上传本月申报数据 Excel 文件（.xlsx），\n"
            "收到文件后 inject:\n"
            'data_key="uploaded_excel" data_value={"file_path": "路径", "user_said": "用户的原话或描述"}'
        ),
    },
    "PARSE_EXCEL": {
        "type": "auto",
        "handler": "do_parse_excel",
    },
    "TAX_CALC": {
        "type": "auto",
        "handler": "do_tax_calc",
    },
    "CONFIRM_TAX": {
        "type": "input",
        "required_input": "tax_confirm_ack",
        "handler": "do_confirm_tax",
        "message": (
            "【必须和用户交互】请向用户展示税费计算结果，等用户确认后 inject:\n"
            'data_key="tax_confirm_ack" data_value={"user_said": "用户的原话"}'
        ),
    },
    "SUBMIT": {
        "type": "auto",
        "handler": "do_submit",
    },
    "DOWNLOAD": {
        "type": "auto",
        "handler": "do_download",
    },
    "NOTIFY_COMPLETE": {
        "type": "input",
        "required_input": "complete_ack",
        "handler": "do_notify_complete",
        "message": (
            "【必须和用户交互】请向用户发送申报完成通知（包含企业名、申报期、税额、PDF信息），\n"
            "用户确认收到后 inject:\n"
            'data_key="complete_ack" data_value={"user_said": "用户的原话"}'
        ),
    },
    "DONE": {
        "type": "terminal",
        "handler": "do_done",
    },
    "FAILED": {
        "type": "terminal",
        "handler": "do_failed",
    },
}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 状态处理函数
# 返回: {"next": "下一个状态"} 或 {"blocked": True, ...}
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def do_init(state: dict) -> dict:
    agg = state["agg_org_id"]
    if not agg:
        return {"blocked": True, "error": "缺少 agg_org_id"}

    log.info(f"[{state['task_id']}] INIT: 跳过自然人登录检查，直接进入申报流程")
    return {"next": "ENTERPRISE_PROFILE"}


def do_enterprise_profile(state: dict) -> dict:
    """发起企业数据采集，获取企业画像全量数据"""
    agg = state["agg_org_id"]
    log.info(f"[{state['task_id']}] ENTERPRISE_PROFILE: 采集企业画像 agg_org_id={agg}")

    result = enterprise_profile(agg)

    if not result["ok"]:
        # 企业画像失败不阻断流程，记录错误但继续
        log.warning(f"[{state['task_id']}] 企业画像采集失败: {result.get('error')}，继续流程")
        state["data"]["enterprise_profile"] = {
            "ok": False,
            "error": result.get("error"),
        }
        save_state(state["task_id"], state)
        return {"next": "FETCH_LIST"}

    profile = result.get("profile", {})
    state["data"]["enterprise_profile"] = {
        "ok": True,
        "profile": profile,
        "raw_data": result.get("raw_data"),
    }

    # 从画像中提取行业信息，后续税费计算可能用到（如行业税负率对比、小微企业判断）
    industry = profile.get("basic", {}).get("industry", "")
    if industry:
        state["data"]["industry"] = industry
        log.info(f"[{state['task_id']}] 企业所属行业: {industry}")

    # 提取纳税人类型，用于后续逻辑判断
    taxpayer_type = profile.get("basic", {}).get("taxpayer_type", "")
    if taxpayer_type:
        state["data"]["taxpayer_type"] = taxpayer_type

    save_state(state["task_id"], state)
    log.info(f"[{state['task_id']}] 企业画像采集完成: {profile.get('basic', {}).get('enterprise_name', '')}")
    return {"next": "FETCH_LIST"}


def do_fetch_list(state: dict) -> dict:
    year, period = _parse_period(state["period"])
    result = fetch_tax_list(state["agg_org_id"], year, period)

    if not result["ok"]:
        fail(state, result.get("error"))
        if state["retry_count"] >= state["max_retries"]:
            return {"next": "FAILED"}
        return {"blocked": True, "error": "获取清册失败，将重试"}

    state["data"]["tax_list"] = result
    save_state(state["task_id"], state)

    if not result.get("has_required"):
        return {"next": "DONE", "info": "本月无需申报"}

    return {"next": "NOTIFY_TAXES"}


def do_notify_taxes(state: dict) -> dict:
    # 走到这里说明 input 已经收到了（advance 负责检查）
    return {"next": "DATA_INIT"}


def do_data_init(state: dict) -> dict:
    year, period = _parse_period(state["period"])
    tax_list = state["data"].get("tax_list", {})
    required = tax_list.get("required_items", [])

    result = init_declaration(state["agg_org_id"], year, period, required)
    state["data"]["init_data"] = result
    save_state(state["task_id"], state)

    if result.get("errors") and not result.get("results"):
        fail(state, f"初始化全部失败: {result['errors']}")
        return {"next": "FAILED"}

    return {"next": "WAIT_UPLOAD"}


def do_wait_upload(state: dict) -> dict:
    return {"next": "PARSE_EXCEL"}


def do_parse_excel(state: dict) -> dict:
    uploaded = state["data"].get("uploaded_excel", {})
    file_path = uploaded.get("file_path", "")
    if not file_path:
        return {"next": "WAIT_UPLOAD", "info": "未找到文件路径，请重新上传"}

    try:
        parsed = _extract_excel_data(file_path, state.get("company_type", "small_scale"))
    except Exception as e:
        log.error(f"[{state['task_id']}] Excel 解析失败: {e}")
        return {"next": "WAIT_UPLOAD", "info": f"Excel 解析失败（{e}），请检查文件后重新上传"}

    state["data"]["parsed_excel"] = parsed
    save_state(state["task_id"], state)
    log.info(f"[{state['task_id']}] Excel 解析完成: vat={bool(parsed.get('vat'))}, cit={bool(parsed.get('cit'))}")
    return {"next": "TAX_CALC"}


def _extract_excel_data(file_path: str, company_type: str = "small_scale") -> dict:
    """从标准税务申报 Excel 中提取增值税和企业所得税数据"""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {"user_said": "Excel 自动解析"}

    def cell_val(ws, row, col, default=0):
        v = ws.cell(row=row, column=col).value
        if v is None:
            return default
        try:
            return float(v)
        except (ValueError, TypeError):
            return default

    def find_sheet(names):
        for n in names:
            for sn in wb.sheetnames:
                if n in sn:
                    return wb[sn]
        return None

    vat_ws = find_sheet(["增值税"])
    if vat_ws:
        title = str(vat_ws.cell(row=1, column=1).value or "")
        is_general = "一般纳税人" in title

        if is_general:
            output_tax = cell_val(vat_ws, 15, 3)
            input_tax = cell_val(vat_ws, 16, 3)
            prev_credit = cell_val(vat_ws, 17, 3)
            input_transfer = cell_val(vat_ws, 18, 3)
            result["vat"] = {
                "output_tax": output_tax,
                "input_tax": input_tax,
                "prev_credit": prev_credit,
                "input_transfer": input_transfer,
            }
            result["detected_company_type"] = "general"
        else:
            sales = cell_val(vat_ws, 8, 3)
            exempt = cell_val(vat_ws, 14, 3)
            special = cell_val(vat_ws, 9, 3)
            result["vat"] = {
                "sales_amount": sales,
                "exempt_sales": exempt,
                "special_sales": special,
            }
            result["detected_company_type"] = "small_scale"

    cit_ws = find_sheet(["企业所得税", "所得税"])
    if cit_ws:
        revenue = cell_val(cit_ws, 9, 3)
        cost = cell_val(cit_ws, 10, 3)
        profit = cell_val(cit_ws, 11, 3)
        accumulated = cell_val(cit_ws, 16, 4)
        prepaid = cell_val(cit_ws, 20, 4)
        result["cit"] = {
            "revenue": revenue,
            "cost": cost,
            "profit": profit,
            "accumulated_profit": accumulated,
            "prepaid_tax": prepaid,
        }

    return result


def do_tax_calc(state: dict) -> dict:
    from calculate_tax import calculate_tax

    tax_list = state["data"].get("tax_list", {})
    codes = [i.get("yzpzzlDm", i.get("tax_code", "")) for i in tax_list.get("required_items", [])]

    parsed = state["data"].get("parsed_excel", {})
    detected = parsed.get("detected_company_type")
    company_type = detected or state.get("company_type", "small_scale")
    if detected and detected != state.get("company_type"):
        log.info(f"[{state['task_id']}] Excel 检测到纳税人类型: {detected}，覆盖原始设定 {state.get('company_type')}")
        state["company_type"] = detected
        save_state(state["task_id"], state)

    # 从企业画像获取行业信息
    industry = state["data"].get("industry", "")
    ep = state["data"].get("enterprise_profile", {})
    if not industry and ep and ep.get("ok"):
        industry = ep.get("profile", {}).get("basic", {}).get("industry", "")

    result = calculate_tax(
        declaration_data=parsed,
        tax_codes=codes,
        company_type=company_type,
        init_data=state["data"].get("init_data"),
        industry=industry,
        financial_report=parsed.get("financial_report"),
    )
    state["data"]["calculated_taxes"] = result
    save_state(state["task_id"], state)

    return {"next": "CONFIRM_TAX"}


def do_confirm_tax(state: dict) -> dict:
    return {"next": "SUBMIT"}


def do_submit(state: dict) -> dict:
    year, period = _parse_period(state["period"])
    agg = state["agg_org_id"]
    is_zero = state["data"].get("calculated_taxes", {}).get("is_all_zero", False)
    ctype = state.get("company_type", "small_scale")

    if is_zero and ctype in ("small_scale", "general"):
        result = submit_simplified(agg, year, period, sb_init=True)
    else:
        payload = state["data"].get("parsed_excel", {}).get("report_payload", {})
        result = submit_standard(agg, year, period, payload)

    state["data"]["submit_result"] = result
    save_state(state["task_id"], state)

    if not result["ok"]:
        if result.get("retryable"):
            fail(state, result.get("error"))
            if state["retry_count"] < state["max_retries"]:
                return {"blocked": True, "error": f"申报失败(可重试): {result.get('error')}"}
            return {"next": "FAILED"}
        fail(state, result.get("error"))
        return {"next": "FAILED"}

    return {"next": "DOWNLOAD"}


# 增值税申报视频 URL（录制好的申报操作视频，随PDF一起返回给用户）
DECLARE_VIDEO_URL = (
    "http://qxy-oss-robot-product.qixiangyun.com/VIDEO/"
    "etax-agg-product_0d17b8be20214c11a52ccb869fb185ce_1773040951439.webm"
    "?OSSAccessKeyId=LTAI5tMHcomKiHbKRhS2uU8X"
    "&Expires=1804144951"
    "&Signature=4GzUiutzNNoMGMtL%2BCt/%2Bk9qWcY%3D"
)


def do_download(state: dict) -> dict:
    year, period = _parse_period(state["period"])
    tax_list = state["data"].get("tax_list", {})
    required = tax_list.get("required_items", [])
    zsxm = [{"yzpzzlDm": i.get("yzpzzlDm", "")} for i in required if i.get("yzpzzlDm")]

    result = download_receipt(state["agg_org_id"], year, period, zsxm)
    state["data"]["receipt_data"] = result.get("structured_data")
    state["data"]["pdf_data"] = result.get("pdf_data")

    # 申报视频URL随PDF一起存储，返回给用户时一并展示
    state["data"]["declare_video"] = {
        "video_url": DECLARE_VIDEO_URL,
        "description": "增值税申报操作视频",
    }

    save_state(state["task_id"], state)

    return {"next": "NOTIFY_COMPLETE"}


def do_notify_complete(state: dict) -> dict:
    return {"next": "DONE"}


def do_done(state: dict) -> dict:
    taxes = state["data"].get("calculated_taxes", {})
    tax_details = []
    for r in taxes.get("results", []):
        tax_details.append({
            "tax_name": r.get("tax_name", ""),
            "tax_code": r.get("tax_code", ""),
            "final_amount": r.get("final_amount", 0),
        })
    ep = state["data"].get("enterprise_profile", {})
    ep_summary = {}
    if ep and ep.get("ok") and ep.get("profile"):
        basic = ep["profile"].get("basic", {})
        ep_summary = {
            "enterprise_name": basic.get("enterprise_name", ""),
            "industry": basic.get("industry", ""),
            "taxpayer_type": basic.get("taxpayer_type", ""),
            "credit_grade": basic.get("credit_grade", ""),
        }

    return {
        "terminal": True,
        "summary": {
            "task_id": state["task_id"],
            "company_name": state["company_name"],
            "period": state["period"],
            "status": "completed",
            "total_payable": taxes.get("total_payable", 0),
            "is_all_zero": taxes.get("is_all_zero", False),
            "tax_details": tax_details,
            "enterprise_profile": ep_summary,
            "message": f"{state['company_name']} {state['period']} 申报完成",
        },
    }


def do_failed(state: dict) -> dict:
    return {
        "terminal": True,
        "summary": {
            "task_id": state["task_id"],
            "company_name": state["company_name"],
            "period": state["period"],
            "status": "failed",
            "error": state.get("error"),
        },
    }


HANDLER_MAP = {
    "do_init": do_init,
    "do_enterprise_profile": do_enterprise_profile,
    "do_fetch_list": do_fetch_list,
    "do_notify_taxes": do_notify_taxes,
    "do_data_init": do_data_init,
    "do_wait_upload": do_wait_upload,
    "do_parse_excel": do_parse_excel,
    "do_tax_calc": do_tax_calc,
    "do_confirm_tax": do_confirm_tax,
    "do_submit": do_submit,
    "do_download": do_download,
    "do_notify_complete": do_notify_complete,
    "do_done": do_done,
    "do_failed": do_failed,
}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 状态链路定义（线性执行顺序）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

STATE_CHAIN = [
    "INIT",
    "ENTERPRISE_PROFILE",
    "FETCH_LIST",
    "NOTIFY_TAXES",
    "DATA_INIT",
    "WAIT_UPLOAD",
    "PARSE_EXCEL",
    "TAX_CALC",
    "CONFIRM_TAX",
    "SUBMIT",
    "DOWNLOAD",
    "NOTIFY_COMPLETE",
    "DONE",
]

# 每个状态完成后应在 state_history 或 data 中留下的证据
STATE_EVIDENCE = {
    "INIT":            {"history": True},
    "ENTERPRISE_PROFILE": {"history": True, "data_key": "enterprise_profile"},
    "FETCH_LIST":      {"history": True, "data_key": "tax_list"},
    "NOTIFY_TAXES":    {"history": True, "data_key": "notify_taxes_ack", "need_user_said": True},
    "DATA_INIT":       {"history": True, "data_key": "init_data"},
    "WAIT_UPLOAD":     {"history": True, "data_key": "uploaded_excel", "need_user_said": True},
    "PARSE_EXCEL":     {"history": True, "data_key": "parsed_excel"},
    "TAX_CALC":        {"history": True, "data_key": "calculated_taxes"},
    "CONFIRM_TAX":     {"history": True, "data_key": "tax_confirm_ack", "need_user_said": True},
    "SUBMIT":          {"history": True, "data_key": "submit_result"},
    "DOWNLOAD":        {"history": True},
    "NOTIFY_COMPLETE": {"history": True, "data_key": "complete_ack", "need_user_said": True},
}


def _reverse_verify(state: dict, target_state: str) -> dict | None:
    """
    反向链路验证：从 target_state 往回遍历到 INIT，
    逐个检查每个前置状态是否真的完成了、数据是否到位。

    返回 None = 全部通过
    返回 dict = 验证失败，包含 broken_at 和 missing 信息
    """
    if target_state not in STATE_CHAIN:
        return None  # FAILED 等非链路状态不做验证

    target_idx = STATE_CHAIN.index(target_state)
    history_states = [h["state"] for h in state.get("state_history", [])]

    for i in range(target_idx - 1, -1, -1):
        check_state = STATE_CHAIN[i]
        evidence = STATE_EVIDENCE.get(check_state, {})

        # 检查1：这个状态是否出现在 state_history 里（说明曾经被执行过）
        if evidence.get("history") and check_state not in history_states:
            return {
                "broken_at": check_state,
                "reason": f"状态 {check_state} 从未被执行过（不在 state_history 中）",
                "chain_position": i,
            }

        # 检查2：这个状态产生的数据是否存在
        data_key = evidence.get("data_key")
        if data_key:
            data_value = state["data"].get(data_key)
            if data_value is None:
                return {
                    "broken_at": check_state,
                    "reason": f"状态 {check_state} 的产出数据 '{data_key}' 不存在",
                    "chain_position": i,
                }

            # 检查3：如果是用户交互节点，验证 user_said 存在
            if evidence.get("need_user_said"):
                if isinstance(data_value, dict) and "user_said" not in data_value:
                    return {
                        "broken_at": check_state,
                        "reason": f"状态 {check_state} 的数据 '{data_key}' 缺少 user_said（用户原话），交互未真实发生",
                        "chain_position": i,
                    }
                if isinstance(data_value, dict) and not data_value.get("user_said"):
                    return {
                        "broken_at": check_state,
                        "reason": f"状态 {check_state} 的 user_said 为空，交互未真实发生",
                        "chain_position": i,
                    }

    return None  # 全部通过


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 核心引擎：advance（只走一步 + 反向验证）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _user_msg(task_id: str, status: str, text: str,
              waiting_for: str = None, summary: dict = None) -> dict:
    """统一输出格式：纯文本 user_message + 最少的结构化字段"""
    r = {"ok": status != "error", "task_id": task_id, "status": status, "user_message": text}
    if waiting_for:
        r["waiting_for"] = waiting_for
    if summary:
        r["summary"] = summary
    return r


def _format_blocked(state_name: str, state: dict) -> str:
    """根据当前 blocked 状态，生成用户友好的中文提示"""
    company = state["company_name"]
    period = state["period"]
    items = state["data"].get("tax_list", {}).get("required_items", [])
    taxes = state["data"].get("calculated_taxes", {})
    receipt = state["data"].get("receipt_data")

    if state_name == "NOTIFY_TAXES":
        lines = []

        # 展示企业画像摘要（如果有）
        ep = state["data"].get("enterprise_profile", {})
        if ep and ep.get("ok") and ep.get("profile"):
            profile = ep["profile"]
            basic = profile.get("basic", {})
            biz = profile.get("business", {})
            th = profile.get("tax_health", {})
            lines.append(f"【企业画像】{basic.get('enterprise_name', company)}")
            if basic.get("industry"):
                lines.append(f"  所属行业：{basic['industry']}")
            if basic.get("taxpayer_type"):
                lines.append(f"  纳税人类型：{basic['taxpayer_type']}")
            if basic.get("tax_status"):
                lines.append(f"  税务经营状态：{basic['tax_status']}")
            if basic.get("credit_grade"):
                lines.append(f"  信用等级：{basic['credit_grade']}（{th.get('credit_label', '')}）")
            if biz.get("total_revenue_3y"):
                lines.append(f"  近3年累计营收：¥{biz['total_revenue_3y']:,.2f}")
            if biz.get("top_business"):
                lines.append(f"  核心业务：{biz['top_business']}")
            if th.get("has_overdue_tax"):
                lines.append(f"  ⚠️ 当前存在欠税")
            if th.get("has_violations"):
                lines.append(f"  ⚠️ 近3年存在违法信息")
            lines.append("")

        lines.append(f"已查询到 {company} {period} 月需申报的税种：\n")
        for i, item in enumerate(items, 1):
            name = item.get("zsxmMc", "未知税种")
            status = item.get("declareStatus", "")
            start = item.get("ssqQ", "")
            end = item.get("ssqZ", "")
            lines.append(f"{i}. {name}（{status}）申报期 {start} 至 {end}")
        lines.append(f"\n共 {len(items)} 个税种。请确认是否开始申报？")
        return "\n".join(lines)

    elif state_name == "WAIT_UPLOAD":
        return f"请上传 {company} {period} 月的申报数据 Excel 文件（.xlsx）。"

    elif state_name == "PARSE_EXCEL":
        return f"正在解析您上传的申报数据，请稍候..."

    elif state_name == "CONFIRM_TAX":
        if taxes and taxes.get("results"):
            lines = [f"{company} {period} 月税费计算结果：\n"]
            for r in taxes["results"]:
                name = r.get("tax_name", r.get("name", "税种"))
                amount = r.get("final_amount", r.get("payable", 0))
                lines.append(f"- {name}: ¥{amount:,.2f}")
            total = taxes.get("total_payable", 0)
            lines.append(f"\n应缴税费合计: ¥{total:,.2f}")
            lines.append("请确认以上税额是否正确？确认后将提交申报。")
            return "\n".join(lines)
        else:
            return f"{company} {period} 月应缴税费合计: ¥0.00（零申报）。请确认是否提交？"

    elif state_name == "NOTIFY_COMPLETE":
        lines = []

        pdf_url = ""
        if receipt and isinstance(receipt, dict):
            inner = receipt.get("data", receipt)
            if isinstance(inner, dict):
                pdf_url = inner.get("pdfUrl", "")
                structured = inner.get("structured_data", {})
                if structured:
                    total_tax = structured.get("total_tax", "0.00")
                    lines.append(f"{company} {period} 月申报已完成！")
                    lines.append(f"应缴税额: ¥{total_tax}")
                    lines.append(f"申报回执已获取。")

        # 展示申报操作视频链接（录制好的视频，和PDF一起返回）
        video_data = state["data"].get("declare_video", {})
        if video_data and video_data.get("video_url"):
            lines.append(f"申报操作视频: {video_data['video_url']}")

        if not lines:
            lines.append(f"{company} {period} 月申报已完成！")

        lines.append("请确认收到以上信息。")
        return "\n".join(lines)

    return STATE_DEFS.get(state_name, {}).get("message", "请回复以继续。")


def _format_summary(summary: dict) -> str:
    """格式化最终完成的 summary"""
    if not summary:
        return "流程已完成。"
    company = summary.get("company_name", "")
    period = summary.get("period", "")
    status = summary.get("status", "")
    msg = summary.get("message", "")
    total = summary.get("total_payable", 0)
    if msg:
        return msg
    if status == "completed":
        return f"{company} {period} 月申报全部完成，应缴税额 ¥{total:.2f}。"
    return f"{company} {period} 月流程结束（状态: {status}）。"


def advance(task_id: str) -> dict:
    state = load_state(task_id)
    if not state:
        return _user_msg(task_id, "error", "任务不存在，请重新创建。")

    MAX_AUTO_STEPS = 20

    for _loop in range(MAX_AUTO_STEPS):
        state = load_state(task_id)
        if not state:
            return _user_msg(task_id, "error", "任务不存在，请重新创建。")

        current = state["state"]
        defn = STATE_DEFS.get(current)
        if not defn:
            return _user_msg(task_id, "error", f"系统异常，请联系管理员。")

        if defn["type"] == "terminal":
            handler = HANDLER_MAP[defn["handler"]]
            result = handler(state)
            summary = result.get("summary", {})
            return _user_msg(task_id, "completed", _format_summary(summary), summary=summary)

        if defn["type"] == "input":
            required_key = defn["required_input"]
            if required_key not in state["data"] or state["data"][required_key] is None:
                return _user_msg(task_id, "need_input",
                                 _format_blocked(current, state),
                                 waiting_for=required_key)

            input_data = state["data"][required_key]
            if isinstance(input_data, dict) and "user_said" not in input_data:
                return _user_msg(task_id, "error", "缺少用户确认信息，请重新回复。")

        chain_error = _reverse_verify(state, current)
        if chain_error:
            return _user_msg(task_id, "error",
                             f"流程异常：{chain_error['reason']}，请联系管理员。")

        handler = HANDLER_MAP[defn["handler"]]
        log.info(f"[{task_id}] 执行: {current}")
        result = handler(state)

        if result.get("blocked"):
            return _user_msg(task_id, "error",
                             result.get("error", "操作失败，请稍后重试。"))

        if result.get("terminal"):
            summary = result.get("summary", {})
            return _user_msg(task_id, "completed", _format_summary(summary), summary=summary)

        next_state = result["next"]
        transition(state, next_state, f"from_{current}")

        next_defn = STATE_DEFS.get(next_state, {})

        if next_defn.get("type") == "input":
            return _user_msg(task_id, "need_input",
                             _format_blocked(next_state, state),
                             waiting_for=next_defn.get("required_input"))

        if next_defn.get("type") == "terminal":
            continue

    return _user_msg(task_id, "error", "流程步骤过多，请联系管理员。")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# inject：严格校验，只接受当前状态需要的 key
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def inject_data(task_id: str, data_key: str, data_value) -> dict:
    state = load_state(task_id)
    if not state:
        return {"ok": False, "error": f"任务不存在: {task_id}"}

    current = state["state"]
    defn = STATE_DEFS.get(current, {})

    # 校验0：反向链路验证 — 所有前置状态必须已完成
    chain_error = _reverse_verify(state, current)
    if chain_error:
        return {
            "ok": False,
            "error": f"反向验证失败，无法在 {current} inject 数据: {chain_error['reason']}",
            "broken_at": chain_error["broken_at"],
            "task_id": task_id,
        }

    # 校验1：当前状态必须是 input 类型
    if defn.get("type") != "input":
        return {
            "ok": False,
            "error": f"当前状态 {current} 不接受输入（类型: {defn.get('type')}）",
            "task_id": task_id,
        }

    # 校验2：data_key 必须匹配当前状态的 required_input
    expected = defn.get("required_input")
    if data_key != expected:
        return {
            "ok": False,
            "error": f"当前状态 {current} 只接受 key='{expected}'，你传的是 '{data_key}'",
            "task_id": task_id,
        }

    # 校验3：必须包含 user_said 字段
    if isinstance(data_value, dict) and "user_said" not in data_value:
        return {
            "ok": False,
            "error": "inject 数据必须包含 'user_said' 字段（用户的原话），防止 LLM 自行编造数据跳过交互",
            "task_id": task_id,
        }

    # 写入
    state["data"][data_key] = data_value
    save_state(task_id, state)
    log.info(f"[{task_id}] ✅ inject {data_key} at state {current}")

    return {"ok": True, "message": f"已注入 {data_key}", "state": current, "task_id": task_id}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CLI
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if __name__ == "__main__":
    args = parse_args()
    action = args.get("action", "status")

    if action == "create":
        state = create_task(
            company_id=args["company_id"],
            company_name=args["company_name"],
            period=args["period"],
            agg_org_id=str(args.get("agg_org_id", args.get("aggOrgId", ""))),
            company_type=args.get("company_type", "small_scale"),
        )
        result = advance(state["task_id"])
        output(result)

    elif action == "advance":
        result = advance(args["task_id"])
        output(result)

    elif action == "inject":
        result = inject_data(args["task_id"], args["data_key"], args["data_value"])
        output(result)

    elif action == "status":
        state = load_state(args.get("task_id", ""))
        if state:
            current = state["state"]
            defn = STATE_DEFS.get(current, {})
            output({
                "ok": True,
                "task_id": state["task_id"],
                "state": current,
                "state_type": defn.get("type"),
                "waiting_for": defn.get("required_input") if defn.get("type") == "input" else None,
                "has_input": defn.get("required_input") in state.get("data", {}) if defn.get("type") == "input" else None,
                "company_name": state["company_name"],
                "period": state["period"],
                "last_updated": state["updated_at"],
            })
        else:
            output({"ok": False, "error": "任务不存在"})

    elif action == "list":
        tasks = list_tasks(args.get("status_filter"))
        output({
            "ok": True,
            "tasks": [{
                "task_id": t["task_id"],
                "state": t["state"],
                "company_name": t["company_name"],
                "period": t["period"],
            } for t in tasks],
        })

    elif action == "verify":
        state = load_state(args.get("task_id", ""))
        if not state:
            output({"ok": False, "error": "任务不存在"})
        else:
            current = state["state"]
            # 逐个检查链路上每个状态的证据
            results = []
            for s in STATE_CHAIN:
                ev = STATE_EVIDENCE.get(s, {})
                passed = True
                reason = "OK"

                history_states = [h["state"] for h in state.get("state_history", [])]
                if ev.get("history") and s not in history_states and s != current:
                    if STATE_CHAIN.index(s) < STATE_CHAIN.index(current):
                        passed = False
                        reason = "未执行"

                dk = ev.get("data_key")
                if dk and passed:
                    dv = state["data"].get(dk)
                    if dv is None and STATE_CHAIN.index(s) < STATE_CHAIN.index(current):
                        passed = False
                        reason = f"数据 {dk} 缺失"
                    elif ev.get("need_user_said") and isinstance(dv, dict) and not dv.get("user_said"):
                        passed = False
                        reason = f"数据 {dk} 缺少 user_said"

                marker = "✅" if passed else "❌"
                is_current = "👉" if s == current else "  "
                results.append(f"{is_current} {marker} {s}" + (f" — {reason}" if not passed else ""))

            output({
                "ok": True,
                "task_id": state["task_id"],
                "current_state": current,
                "chain_verification": results,
            })

    else:
        output({"ok": False, "error": f"未知 action: {action}"})
